import io
import os
import copy
import openpyxl
from openpyxl.cell.cell import MergedCell
from django.conf import settings
from datetime import datetime

def generar_excel_prefactura(data):
    """
    Genera la prefactura a partir de la plantilla oficial en Excel.
    Usa escritura segura para evitar crashear si la plantilla tiene celdas fusionadas sobre los datos.
    """
    template_path = os.path.join(settings.BASE_DIR, 'apps', 'cotizador', 'templates', 'PLANTILLA_FACTURA.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb['4.0'] if '4.0' in wb.sheetnames else wb.active

    # Función inteligente que evita crashear si la celda está bloqueada por una fusión
    def escribir_seguro(coord, valor):
        if not isinstance(ws[coord], MergedCell):
            ws[coord] = valor

    # 2. Llenar los datos de cabecera usando escritura segura
    escribir_seguro('D2', str(data.get('empresa_nombre', '')).strip())
    escribir_seguro('D4', str(data.get('tipo_comprobante', 'I - INGRESO')).strip())
    escribir_seguro('D6', str(data.get('rfc_receptor', '')).strip())
    escribir_seguro('D8', str(data.get('razon_social', '')).strip())
    escribir_seguro('D9', str(data.get('calle_numero', '')).strip())
    escribir_seguro('D10', str(data.get('colonia', '')).strip())
    escribir_seguro('D11', str(data.get('ciudad', '')).strip())
    escribir_seguro('D12', str(data.get('estado', '')).strip())
    escribir_seguro('D13', str(data.get('codigo_postal', '')).strip())
    escribir_seguro('E14', str(data.get('regimen_fiscal', '601 - Régimen General de Ley Personas Morales')).strip())

    escribir_seguro('G8', str(data.get('fecha_pago', datetime.now().strftime('%d/%m/%Y'))).strip())
    escribir_seguro('G9', str(data.get('moneda', 'MXN - Peso Mexicano')).strip())
    escribir_seguro('G10', str(data.get('tipo_cambio', '1')).strip() if 'USD' in data.get('moneda', '') else '')
    escribir_seguro('G11', str(data.get('forma_pago', '03 - TRANSFERENCIA ELECTRÓNICA DE FONDOS')).strip())
    escribir_seguro('G12', str(data.get('metodo_pago', 'PUE - Pago en una sola exhibición')).strip())
    escribir_seguro('G13', str(data.get('uso_cfdi', 'G03 - GASTOS EN GENERAL')).strip())

    # 3. Procesar las partidas dinámicamente
    partidas = data.get('partidas', [])
    if not partidas:
        partidas = [{}]
        
    num_partidas = len(partidas)
    
    if num_partidas > 1:
        ws.insert_rows(21, amount=num_partidas - 1)
        for r in range(21, 20 + num_partidas):
            for c in range(2, 12):
                source_cell = ws.cell(row=20, column=c)
                target_cell = ws.cell(row=r, column=c)
                
                if source_cell.has_style:
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.alignment = copy.copy(source_cell.alignment)

    subtotal_acumulado = 0.0
    impuestos_acumulados = 0.0

    thick = openpyxl.styles.Side(style='medium', color='000000')
    thin = openpyxl.styles.Side(style='thin', color='000000')

    for idx, p in enumerate(partidas):
        row_idx = 20 + idx
        cant = float(p.get('cantidad', 1) or 1)
        val_unit = float(p.get('valor_unitario', 0) or 0)
        tasa_iva = float(p.get('tasa_iva', 0.16) or 0.16)
        
        importe_partida = cant * val_unit
        impuesto_partida = importe_partida * tasa_iva
        
        subtotal_acumulado += importe_partida
        impuestos_acumulados += impuesto_partida
        
        ws.cell(row=row_idx, column=2, value=idx+1)
        ws.cell(row=row_idx, column=3, value=str(p.get('clave_prod', '')))
        ws.cell(row=row_idx, column=4, value=cant)
        ws.cell(row=row_idx, column=5, value=str(p.get('clave_unidad', '')))
        ws.cell(row=row_idx, column=6, value=str(p.get('unidad', 'SERVICIO')))
        ws.cell(row=row_idx, column=7, value=str(p.get('descripcion', '')))
        ws.cell(row=row_idx, column=8, value=val_unit).number_format = '$ #,##0.00'
        ws.cell(row=row_idx, column=9, value=str(p.get('impuesto_label', '002 - IVA')))
        ws.cell(row=row_idx, column=10, value=impuesto_partida).number_format = '$ #,##0.00'
        ws.cell(row=row_idx, column=11, value=importe_partida).number_format = '$ #,##0.00'

        is_last = (idx == num_partidas - 1)
        for c in range(2, 12):
            cell = ws.cell(row=row_idx, column=c)
            old = cell.border
            if old:
                cell.border = openpyxl.styles.Border(
                    left=old.left, right=old.right, top=old.top,
                    bottom=thick if is_last else thin
                )

    # 4. Actualizar Totales
    totales_row_start = 22 + max(0, num_partidas - 1)
    
    escribir_seguro(f'K{totales_row_start}', subtotal_acumulado)
    escribir_seguro(f'K{totales_row_start+1}', impuestos_acumulados)
    escribir_seguro(f'K{totales_row_start+2}', subtotal_acumulado + impuestos_acumulados)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, {
        'subtotal': subtotal_acumulado,
        'impuestos_trasladados': impuestos_acumulados,
        'total': subtotal_acumulado + impuestos_acumulados
    }
