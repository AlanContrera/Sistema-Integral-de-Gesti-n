import os
import openpyxl
from django.core.management.base import BaseCommand
from apps.reclutamiento.models import CategoriaPreguntas, PlantillaPregunta, Estado, Municipio
from django.conf import settings

class Command(BaseCommand):
    help = 'Importa el catálogo de puestos y factores de ubicación'

    def handle(self, *args, **kwargs):
        ruta_excel = os.path.join(settings.BASE_DIR, 'media', 'IACI INGENIERO DE IMPLEMENTACION Y PROYECTO RECLUTAMIENTO BUENO edITABLE.xlsx')
        
        if not os.path.exists(ruta_excel):
            self.stdout.write(self.style.ERROR(f'No se encontró el archivo en: {ruta_excel}'))
            return

        self.stdout.write('Leyendo el Excel (esto puede tardar unos segundos)...')
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        
        if 'Catálogos' not in wb.sheetnames:
            self.stdout.write(self.style.ERROR('No se encontró la pestaña "Catálogos".'))
            return
            
        ws = wb['Catálogos']
        
        # 1. Importar factores de ubicación (Columnas BA: Estado, BB: Municipio, BC: Factor)
        self.stdout.write('Importando factores de ubicación...')
        municipios_actualizados = 0
        for row in ws.iter_rows(min_row=2, min_col=53, max_col=55, values_only=True): # BA=53, BB=54, BC=55
            estado_nombre = row[0]
            municipio_nombre = row[1]
            factor_val = row[2]
            
            if not estado_nombre or not municipio_nombre or factor_val is None:
                continue
                
            estado_nombre = str(estado_nombre).strip()
            municipio_nombre = str(municipio_nombre).strip()
            
            try:
                factor = float(factor_val)
            except:
                factor = 1.0
                
            estado_obj, _ = Estado.objects.get_or_create(nombre=estado_nombre)
            mun_obj, _ = Municipio.objects.get_or_create(estado=estado_obj, nombre=municipio_nombre)
            
            mun_obj.factor_ubicacion = factor
            mun_obj.save()
            municipios_actualizados += 1

        # 2. Importar Puestos y desglosar viñetas
        self.stdout.write('Importando puestos y desglosando viñetas...')
        categorias_creadas = 0
        preguntas_creadas = 0
        
        def procesar_vinetas(texto):
            if not texto:
                return []
            lineas = str(texto).split('\n')
            vinetas = [l.strip().lstrip('•').strip() for l in lineas if l.strip()]
            return [v for v in vinetas if v]

        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            nombre_puesto = row[0]
            if not nombre_puesto:
                continue
                
            nombre_puesto = str(nombre_puesto).strip()
            sueldo_base_val = row[1]
            
            try:
                sueldo_base = float(sueldo_base_val) if sueldo_base_val else 0.0
            except:
                sueldo_base = 0.0
                
            funciones = procesar_vinetas(row[2])
            responsabilidades = procesar_vinetas(row[3])
            comp_tecnicas = procesar_vinetas(row[4])
            comp_blandas = procesar_vinetas(row[5])
            kpis = procesar_vinetas(row[6])

            categoria, created = CategoriaPreguntas.objects.get_or_create(
                nombre=nombre_puesto,
                defaults={
                    'descripcion': 'Importado automáticamente desde Excel',
                    'sueldo_promedio_base': sueldo_base
                }
            )
            
            if not created:
                categoria.sueldo_promedio_base = sueldo_base
                categoria.save()
                categoria.preguntas.all().delete()
            else:
                categorias_creadas += 1
                
            orden = 1
            
            # Crear pregunta por cada viñeta
            for f in funciones:
                PlantillaPregunta.objects.create(categoria=categoria, rubro='Funciones principales', pregunta=f, criterio_evaluacion='Debe explicar experiencia práctica.', orden=orden)
                orden += 1
                preguntas_creadas += 1
                
            for r in responsabilidades:
                PlantillaPregunta.objects.create(categoria=categoria, rubro='Responsabilidades críticas', pregunta=r, criterio_evaluacion='Debe demostrar capacidad de dar resultados.', orden=orden)
                orden += 1
                preguntas_creadas += 1
                
            for ct in comp_tecnicas:
                PlantillaPregunta.objects.create(categoria=categoria, rubro='Competencias técnicas', pregunta=ct, criterio_evaluacion='Debe explicar uso práctico y nivel de dominio.', orden=orden)
                orden += 1
                preguntas_creadas += 1
                
            for cb in comp_blandas:
                PlantillaPregunta.objects.create(categoria=categoria, rubro='Competencias blandas', pregunta=cb, criterio_evaluacion='Debe dar un ejemplo tipo STAR.', orden=orden)
                orden += 1
                preguntas_creadas += 1
                
            for k in kpis:
                PlantillaPregunta.objects.create(categoria=categoria, rubro='Factores clave de éxito', pregunta=k, criterio_evaluacion='Debe mostrar evidencia e indicadores de éxito.', orden=orden)
                orden += 1
                preguntas_creadas += 1
                
        self.stdout.write(self.style.SUCCESS(f'¡Éxito! {municipios_actualizados} municipios actualizados, {categorias_creadas} puestos creados y {preguntas_creadas} viñetas dinámicas generadas.'))
