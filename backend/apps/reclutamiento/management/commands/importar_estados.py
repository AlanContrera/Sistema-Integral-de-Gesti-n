import os
import re
import openpyxl
from django.core.management.base import BaseCommand
from apps.reclutamiento.models import Estado, Municipio
from django.conf import settings

class Command(BaseCommand):
    help = 'Importa el catálogo de Estados y Municipios limpiando números'

    def handle(self, *args, **kwargs):
        ruta_excel = os.path.join(settings.BASE_DIR, 'media', 'IACI INGENIERO DE IMPLEMENTACION Y PROYECTO RECLUTAMIENTO BUENO edITABLE.xlsx')
        
        if not os.path.exists(ruta_excel):
            self.stdout.write(self.style.ERROR(f'No se encontró el archivo: {ruta_excel}'))
            return

        self.stdout.write('Leyendo el Excel...')
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        
        nombre_hoja = 'Estados y Municipios'
        if nombre_hoja not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f'No se encontró la pestaña "{nombre_hoja}"'))
            return
            
        ws = wb[nombre_hoja]
        estados_creados = 0
        municipios_creados = 0

        # Función para quitar números, puntos o guiones al inicio (ej: "01. Aguascalientes" -> "Aguascalientes")
        def limpiar_nombre(texto):
            if not texto: return ""
            return re.sub(r'^[\d\.\-\s]+', '', str(texto)).strip()

        # Ojo aquí: max_col ahora es 3
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            nombre_estado_sucio = row[0]  # Columna A
            # Brincamos la Columna B (row[1]) porque son puros números
            nombre_municipio_sucio = row[2]  # Columna C

            
            if not nombre_estado_sucio:
                continue
                
            nombre_estado = limpiar_nombre(nombre_estado_sucio)
            nombre_municipio = limpiar_nombre(nombre_municipio_sucio)
            
            if not nombre_estado:
                continue

            estado_obj, created_est = Estado.objects.get_or_create(nombre=nombre_estado)
            if created_est:
                estados_creados += 1
                
            if nombre_municipio:
                mun_obj, created_mun = Municipio.objects.get_or_create(
                    estado=estado_obj,
                    nombre=nombre_municipio
                )
                if created_mun:
                    municipios_creados += 1
                
        self.stdout.write(self.style.SUCCESS(f'¡Limpios e Importados! {estados_creados} estados y {municipios_creados} municipios.'))
