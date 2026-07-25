import pandas as pd
import openpyxl
import json

excel_path = '/home/sistemas/Proyectos/App_Facturacion/backend/media/IACI INGENIERO DE IMPLEMENTACION Y PROYECTO RECLUTAMIENTO BUENO edITABLE.xlsx'

# Abrir con openpyxl para ver fórmulas, validaciones, y celdas combinadas
wb = openpyxl.load_workbook(excel_path, data_only=False)

print("=" * 80)
print("ANÁLISIS COMPLETO DEL EXCEL DE RECLUTAMIENTO")
print("=" * 80)
print(f"\nTotal de hojas: {len(wb.sheetnames)}")
print(f"Hojas: {wb.sheetnames}\n")

for idx, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"HOJA {idx+1}: '{sheet_name}'")
    print(f"Dimensiones: {ws.dimensions}")
    print(f"Filas usadas: {ws.max_row}, Columnas usadas: {ws.max_column}")
    print(f"{'='*80}")
    
    # Mostrar TODAS las celdas con contenido
    print("\n--- CONTENIDO CELDA POR CELDA ---")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell_ref = f"{cell.column_letter}{cell.row}"
                val = cell.value
                extra = ""
                if cell.data_type == 'f':
                    extra = f" [FÓRMULA]"
                print(f"  {cell_ref}: {val}{extra}")
    
    # Celdas combinadas
    if ws.merged_cells.ranges:
        print(f"\n--- CELDAS COMBINADAS ---")
        for merged in ws.merged_cells.ranges:
            print(f"  {merged}")
    
    # Validaciones de datos (LISTAS DESPLEGABLES)
    if ws.data_validations and ws.data_validations.dataValidation:
        print(f"\n--- VALIDACIONES DE DATOS (LISTAS DESPLEGABLES) ---")
        for dv in ws.data_validations.dataValidation:
            print(f"  Tipo: {dv.type}")
            print(f"  Celdas: {dv.sqref}")
            print(f"  Fórmula1: {dv.formula1}")
            if dv.formula2:
                print(f"  Fórmula2: {dv.formula2}")
            print(f"  Allow blank: {dv.allow_blank}")
            print()
    
    print(f"\n{'─'*40}")

print("\n\n¡ANÁLISIS COMPLETO!")
