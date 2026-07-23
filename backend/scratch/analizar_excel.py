import openpyxl

ruta = '/home/sistemas/Proyectos/App_Facturacion/backend/media/IACI INGENIERO DE IMPLEMENTACION Y PROYECTO RECLUTAMIENTO BUENO edITABLE.xlsx'
wb = openpyxl.load_workbook(ruta, data_only=False)

# Ver el Catálogo completo - ahí están las preguntas pre-cargadas
for nombre in ['Catálogos', 'Perfilador Reclutamiento']:
    ws = wb[nombre]
    print(f'\n\n======= HOJA: {nombre} ({ws.max_row}f x {ws.max_column}c) =======')
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        fila = []
        for cell in row:
            val = cell.value
            if val is not None:
                texto = str(val)[:120]
                fila.append(f'[{cell.column_letter}{cell.row}: {texto}]')
        if fila:
            print('  '.join(fila))
            count += 1
        if count >= 60:
            print('...(truncado)')
            break
